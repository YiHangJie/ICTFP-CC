import os
import pandas as pd
from openpyxl import load_workbook

def save_to_excel(df, file_path, sheet_name):

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)

        # 设置第一个工作表为可见
        workbook.active = 0

        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
            workbook.save(file_path)
        
        with pd.ExcelWriter(file_path, engine='openpyxl', mode="a") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
    else:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)